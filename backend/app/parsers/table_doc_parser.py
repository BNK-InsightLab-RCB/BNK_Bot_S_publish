"""Table definition document parser."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List

from backend.app.parsers.base import BaseParser, KnowledgeDocument, SourceFile
from backend.app.utils.text import normalize_ws, unique_keep_order


class TableDocParser(BaseParser):
    """Parse Markdown, CSV, and xlsx table-definition documents."""

    supported_extensions = [".md", ".csv", ".xlsx"]

    def parse(self, source: SourceFile) -> List[KnowledgeDocument]:
        if source.path.suffix.lower() == ".csv":
            return [self._parse_csv(source)]
        if source.path.suffix.lower() == ".xlsx":
            return [self._parse_xlsx(source)]
        if source.path.stem.startswith("incident_"):
            return [self._parse_incident(source)]
        return [self._parse_markdown(source)]

    def _parse_markdown(self, source: SourceFile) -> KnowledgeDocument:
        lines = source.text.splitlines()
        heading = next((line.lstrip("# ").strip() for line in lines if line.startswith("#")), source.path.stem)
        table_name = _guess_table_name(source.path, heading)
        columns = []
        for line in lines:
            if "|" not in line or "---" in line:
                continue
            cells = [normalize_ws(cell) for cell in line.strip("|").split("|")]
            if cells and cells[0] and cells[0].upper() not in {"COLUMN", "컬럼"}:
                columns.append(cells[0].upper())
        return KnowledgeDocument(
            doc_type="table_definition",
            title=f"{table_name} table definition",
            summary=f"{table_name} 테이블 정의서",
            source_path=str(source.path),
            tables=[table_name],
            columns=unique_keep_order(columns),
            branch_guide=f"{table_name} 관련 업무 데이터 정의를 확인할 수 있습니다.",
            it_guide=f"{table_name} columns: {', '.join(unique_keep_order(columns))}",
            code_text=source.text,
            end_line=max(1, len(lines)),
        )

    def _parse_csv(self, source: SourceFile) -> KnowledgeDocument:
        rows = list(csv.DictReader(source.text.splitlines()))
        table_name = _guess_table_name(source.path, source.path.stem)
        columns = [row.get("column") or row.get("COLUMN") or row.get("컬럼") or "" for row in rows]
        return KnowledgeDocument(
            doc_type="table_definition",
            title=f"{table_name} table definition",
            summary=f"{table_name} 테이블 정의서",
            source_path=str(source.path),
            tables=[table_name],
            columns=unique_keep_order(columns),
            code_text=source.text,
        )

    def _parse_xlsx(self, source: SourceFile) -> KnowledgeDocument:
        table_name = _guess_table_name(source.path, source.path.stem)
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(source.path, read_only=True, data_only=True)
            sheet = workbook.active
            columns = [str(row[0].value) for row in sheet.iter_rows(min_row=2) if row and row[0].value]
        except Exception:
            columns = []
        return KnowledgeDocument(
            doc_type="table_definition",
            title=f"{table_name} table definition",
            summary=f"{table_name} 테이블 정의서",
            source_path=str(source.path),
            tables=[table_name],
            columns=unique_keep_order(columns),
        )

    def _parse_incident(self, source: SourceFile) -> KnowledgeDocument:
        lines = source.text.splitlines()
        heading = next((line.lstrip("# ").strip() for line in lines if line.startswith("#")), source.path.stem)
        messages = [line.split(":", 1)[1].strip() for line in lines if line.lower().startswith("- error:")]
        return KnowledgeDocument(
            doc_type="incident",
            title=heading,
            summary=normalize_ws(" ".join(line.strip("#- ") for line in lines if line.strip()))[:500],
            source_path=str(source.path),
            screen_name="고객조회" if "고객조회" in source.text else None,
            error_messages=messages,
            business_rules=["과거 동일 오류 이력이 있으면 권한과 고객 상태를 우선 확인한다."],
            branch_guide="동일 권한 오류가 반복되면 권한 보유 여부와 발생 시각을 IT부서에 전달해야 합니다.",
            it_guide="과거 고객 저장 권한 오류 이력",
            code_text=source.text,
            end_line=max(1, len(lines)),
        )


def _guess_table_name(path: Path, heading: str) -> str:
    for token in [path.stem, heading]:
        upper = token.upper()
        if "TB_" in upper:
            return upper[upper.index("TB_") :].split()[0].replace("_TABLE", "")
    return path.stem.upper().replace("_TABLE", "")
